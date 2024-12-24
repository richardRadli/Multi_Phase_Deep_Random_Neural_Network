import colorlog
import gc
import json
import jsonschema
import logging
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
import pandas as pd
import seaborn as sns
import time
import torch
import sys

from datetime import datetime
from functools import wraps
from jsonschema import validate
from nn.dataloaders.npz_dataloader import NpzDataset
from openpyxl.styles import PatternFill
from torch.utils.data import DataLoader
from typing import Any, Callable, List, Optional, Tuple, Union


def average_columns_in_excel(filename: str) -> None:
    """
    Calculates the average of each numeric column for each sheet in an Excel file and appends these averages to the end
    of each sheet.

    Args:
        filename (str): The path to the Excel file.

    Returns:
        None: The function modifies the Excel file in-place.
    """

    excel_file = pd.ExcelFile(filename)

    results = {}

    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(filename, sheet_name=sheet_name)
        column_averages = df.mean(numeric_only=True)
        results[sheet_name] = column_averages

    workbook = openpyxl.load_workbook(filename)

    for sheet_name, avg in results.items():
        sheet = workbook[sheet_name]
        first_empty_row = sheet.max_row + 1

        for col_num, (col_name, value) in enumerate(avg.items(), start=1):
            sheet.cell(row=first_empty_row, column=col_num, value=value)

        fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")
        for cell in sheet[first_empty_row]:
            cell.fill = fill

    workbook.save(filename)


def calc_exp_neurons(total_neurons: int, n_layers: int) -> List[int]:
    """
    Distributes a total number of neurons across a specified number of layers
    using an exponential growth distribution.

    Args:
        total_neurons (int): The total number of neurons to distribute.
        n_layers (int): The number of layers across which to distribute the neurons.

    Returns:
        List[int]: A list of integers representing the number of neurons in each layer,
                   sorted in descending order.
    """

    geom_mean = total_neurons ** (1 / n_layers)
    neurons_per_layer = [int(geom_mean ** i) for i in range(n_layers)]
    neurons_per_layer[0] += total_neurons - sum(neurons_per_layer)

    if neurons_per_layer[-1] <= 1:
        neurons_per_layer[-2] += neurons_per_layer[-1] - 1
        neurons_per_layer.pop()

    return sorted(neurons_per_layer, reverse=True)


def create_dir(root_dir: str, method: str) -> str:
    """
    Creates a new directory with a timestamp and a specified method as its name
    within a given root directory.

    Args:
        root_dir (str): The path to the root directory where the new directory will be created.
        method (str): The method name or identifier to include in the new directory's name.

    Returns:
        str: The path to the newly created directory.
    """

    timestamp = create_timestamp()
    output_dir = os.path.join(root_dir, f"{timestamp}_{method}")
    os.makedirs(output_dir, exist_ok=True)

    return output_dir


def create_timestamp() -> str:
    """
    Creates a timestamp in the format of '%Y-%m-%d_%H-%M-%S', representing the current date and time.

    Returns:
        Timestamp string in the format of '%Y-%m-%d_%H-%M-%S'.
    """

    return datetime.now().strftime('%Y-%m-%d_%H-%M-%S')


def create_train_valid_test_datasets(file_path, batch_size=None) -> Tuple[DataLoader, DataLoader, DataLoader]:
    """
    Creates DataLoader instances for training, validation, and testing datasets from a given file path.

    Args:
        file_path (str): The path to the file from which the datasets will be created.
        batch_size (int): The batch size to use for training. If None, the default whole dataset will be used.

    Returns:
        tuple: A tuple containing three DataLoader instances:
            - `train_loader`: DataLoader for the training dataset.
            - `valid_loader`: DataLoader for the validation dataset.
            - `test_loader`: DataLoader for the testing dataset.
    """

    train_dataset = NpzDataset(file_path, operation="train")
    valid_dataset = NpzDataset(file_path, operation="valid")
    test_dataset = NpzDataset(file_path, operation="test")

    train_loader = (
        DataLoader(
            dataset=train_dataset, batch_size=len(train_dataset) if batch_size is None else batch_size, shuffle=False
        )
    )
    valid_loader = (
        DataLoader(
            dataset=valid_dataset, batch_size=len(valid_dataset) if batch_size is None else batch_size, shuffle=False
        )
    )
    test_loader = (
        DataLoader(
            dataset=test_dataset, batch_size=len(test_dataset) if batch_size is None else batch_size, shuffle=False
        )
    )

    logging.info(f"Size of train dataset: {len(train_dataset)}, Size of test dataset: {len(test_dataset)}")

    return train_loader, valid_loader, test_loader


def device_selector(preferred_device: str) -> torch.device:
    """
    Provides information about the currently available GPUs and returns a torch device for training and inference.

    Args:
        preferred_device: A torch device for either "cuda" or "cpu".

    Returns:
        torch.device: A torch.device object representing the selected device for training and inference.
    """

    if preferred_device not in ["cuda", "cpu"]:
        logging.warning("Preferred device is not valid. Using CPU instead.")
        return torch.device("cpu")

    if preferred_device == "cuda" and torch.cuda.is_available():
        cuda_info = {
            'CUDA Available': [torch.cuda.is_available()],
            'CUDA Device Count': [torch.cuda.device_count()],
            'Current CUDA Device': [torch.cuda.current_device()],
            'CUDA Device Name': [torch.cuda.get_device_name(0)]
        }

        df = pd.DataFrame(cuda_info)
        logging.info(df)
        return torch.device("cuda")

    if preferred_device in ["cuda"] and not torch.cuda.is_available():
        logging.info("Only CPU is available!")
        return torch.device("cpu")

    if preferred_device == "cpu":
        logging.info("Selected CPU device")
        return torch.device("cpu")


def exponential_neurons(num_of_layers, num_of_neurons, decay_rate=0.5):
    if num_of_layers <= 0:
        raise ValueError("Number of layers must be greater than zero.")
    if num_of_neurons <= 0:
        raise ValueError("Number of neurons must be greater than zero.")

    layers = np.arange(num_of_layers)
    neuron_distribution = np.exp(-decay_rate * layers)

    neuron_distribution /= neuron_distribution.sum()
    neuron_distribution *= num_of_neurons

    neuron_distribution = np.round(neuron_distribution).astype(int)

    while neuron_distribution.sum() < num_of_neurons:
        for i in range(len(neuron_distribution)):
            if neuron_distribution[i] > 0:
                neuron_distribution[i] += 1
                if neuron_distribution.sum() >= num_of_neurons:
                    break

    return neuron_distribution.tolist()


def find_latest_file_in_latest_directory(path: str) -> str:
    """
    Finds and returns the path of the latest file in the most recently modified directory within a given path.

    Args:
        path (str): The path to the parent directory containing subdirectories with files.

    Returns:
        str: The path to the latest file in the most recently modified directory.

    Raises:
        ValueError: If no directories are found within the given path or if no files are found in the latest directory.
    """

    dirs = [os.path.join(path, d) for d in os.listdir(path) if os.path.isdir(os.path.join(path, d))]

    if not dirs:
        raise ValueError(f"No directories found in {path}")

    dirs.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    latest_dir = dirs[0]
    files = [os.path.join(latest_dir, f) for f in os.listdir(latest_dir) if
             os.path.isfile(os.path.join(latest_dir, f))]

    if not files:
        raise ValueError(f"No files found in {latest_dir}")

    files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    latest_file = files[0]
    logging.info(f"Latest file found: {latest_file}")

    return latest_file


def get_num_of_neurons(cfg: dict, method: str) -> list:
    """
    Retrieves the number of neurons for a given method from a configuration dictionary.

    Args:
        cfg (dict): A dictionary containing configuration settings, including neuron counts for different methods.
        method (str): The method for which the number of neurons is requested (e.g., "BASE", "EXP_ORT").

    Returns:
        int: The number of neurons associated with the specified method.

    Raises:
        KeyError: If the provided method is not found in the configuration.
    """

    num_neurons = {
        "BASE": cfg.get("eq_neurons"),
        "EXP_ORT": cfg.get("exp_neurons"),
        "EXP_ORT_C": cfg.get("exp_neurons"),
    }
    return num_neurons[method]


def insert_data_to_excel(filename: str, dataset_name: str, row: int, data: list) -> None:
    """
    Inserts a row of data into a specific sheet of an Excel workbook, creating the sheet if it doesn't exist.

    Args:
        filename (str): The path to the Excel file where data will be inserted.
        dataset_name (str): The name of the sheet in which to insert the data.
        row (int): The row number in the sheet where the data should be inserted.
        data (List[List[Union[str, int, float]]]): A list containing rows of data to be inserted into the specified row.
            Each row is a list of values, which can be strings, integers, or floats.

    Returns:
        None: The function modifies the Excel file in-place.
    """

    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    if dataset_name not in workbook.sheetnames:
        workbook.create_sheet(dataset_name)

    sheet = workbook[dataset_name]

    values = ["train acc",
              "test acc",
              "train precision",
              "test precision",
              "train recall",
              "test recall",
              "train f1",
              "test f1",
              "training time"]

    for col, value in enumerate(values, start=1):
        sheet.cell(row=1, column=col, value=value)

    for col, value in enumerate(data[0], start=1):
        sheet.cell(row=row, column=col, value=str(value))

    if "Sheet" in workbook.sheetnames:
        del workbook['Sheet']

    workbook.save(filename)


def load_config_json(json_schema_filename: str, json_filename: str):
    """
    Loads and validates a JSON configuration file against a JSON schema, flattens and processes the configuration,
    and returns the processed configuration as a dictionary.

    Args:
        json_schema_filename (str): The path to the JSON schema file used for validation.
        json_filename (str): The path to the JSON configuration file to be validated and processed.

    Returns:
        Dict[str, Any]: A dictionary containing the processed configuration. Keys are configuration parameters,
                        and values are their corresponding values.

    Raises:
        jsonschema.exceptions.ValidationError: If the JSON data does not conform to the schema.
    """

    with open(json_schema_filename, "r") as schema_file:
        schema = json.load(schema_file)

    with open(json_filename, "r") as config_file:
        config = json.load(config_file)

    try:
        validate(config, schema)
        logging.info("JSON data is valid.")

        pd.set_option('display.max_colwidth', None)
        pd.set_option('display.max_rows', None)

        flattened_config = {}

        simple_config = {k: v for k, v in config.items() if not isinstance(v, dict)}
        nested_config = {k: v for k, v in config.items() if isinstance(v, dict)}

        dataset_name = simple_config['dataset_name']

        if 'hyperparamtuning' in nested_config:
            flattened_config['hyperparamtuning'] = nested_config['hyperparamtuning']

        if 'optimization' in nested_config:
            flattened_config['optimization'] = nested_config['optimization']

        for key, value in nested_config.items():
            if key != 'hyperparamtuning' and dataset_name in value:
                flattened_config[key] = value[dataset_name]

        full_config = {**simple_config, **flattened_config}
        df = pd.DataFrame.from_dict(full_config, orient='index', columns=['Value'])

        logging.info("Config DataFrame:\n" + df.to_string())

        return full_config
    except jsonschema.exceptions.ValidationError as err:
        logging.error(f"JSON data is invalid: {err}")


def log_to_excel(execution_time, accuracy, file_path):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Log'
        sheet.append(["Total Execution Time", "Accuracy"])
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

    total_execution_time = sum(execution_time)
    sheet.append([total_execution_time, accuracy])
    workbook.save(file_path)


def measure_execution_time(func: Callable) -> Callable:
    """
    Decorator to measure the execution time of a function.

    Args:
        func (Callable): The function to be decorated.

    Returns:
        Callable: The decorated function.
    """

    @wraps(func)
    def wrapper(*args, **kwargs) -> Any:
        """
        Wrapper function to measure execution time.

        Args:
            *args: Positional arguments passed to the function.
            **kwargs: Keyword arguments passed to the function.

        Returns:
            Any: The result of the function.
        """

        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        wrapper.execution_time = end_time - start_time
        logging.info(f"Execution time of {func.__name__}: {wrapper.execution_time:.4f} seconds")
        return result

    wrapper.execution_time = None
    return wrapper


def plot_confusion_matrices_fcnn_helm(cm_train: np.ndarray, cm_test: np.ndarray, class_labels: List[str],
                                      dataset_name: str, path_to_plot: str, index: int) -> None:
    """
    Plots two confusion matrices (training and testing) as heatmaps side by side for a given dataset and operation.

    Args:
        cm_train (np.ndarray): A 2D NumPy array representing the training confusion matrix.
        cm_test (np.ndarray): A 2D NumPy array representing the testing confusion matrix.
        class_labels (List[str]): A list of class labels for the confusion matrix axes.
        dataset_name (str): The name of the dataset for which the confusion matrices are plotted.
        path_to_plot (str): The path to the plot file to be saved.
        index (int): The index of the current test iteration.

    Returns:
        None: The function displays the plot and does not return any value.
    """

    filename = os.path.join(path_to_plot, f"{dataset_name}_{index}.png")

    fig, axes = plt.subplots(1, 2, figsize=(24, 8))

    sns.heatmap(cm_train, annot=True, fmt=".0f", cmap="Blues",
                xticklabels=class_labels, yticklabels=class_labels, ax=axes[0])
    axes[0].set_xlabel("Predicted labels")
    axes[0].set_ylabel("Actual labels")
    axes[0].set_title(f"Training Confusion Matrix of {dataset_name} on the train set")

    sns.heatmap(cm_test, annot=True, fmt=".0f", cmap="Blues",
                xticklabels=class_labels, yticklabels=class_labels, ax=axes[1])
    axes[1].set_xlabel("Predicted labels")
    axes[1].set_ylabel("Actual labels")
    axes[1].set_title(f"Testing Confusion Matrix of {dataset_name} on the test set")

    plt.tight_layout()
    plt.savefig(filename, dpi=300)
    plt.close()
    gc.collect()


def plot_confusion_matrix_mpdrnn(cm: np.ndarray, path_to_plot: str, name_of_dataset: str, operation: str, method: str,
                                 labels=None) -> None:
    """
    Plots multiple confusion matrices side by side and saves the plot as a PNG file.

    Args:
        cm (List[np.ndarray]): A list of 2D NumPy arrays representing confusion matrices to be plotted.
        path_to_plot (str): The directory path where the plot image will be saved.
        name_of_dataset (str): The name of the dataset for which the confusion matrices are plotted.
        operation (str): A string indicating the operation or task (e.g., "training", "validation", "testing").
        method (str): A string indicating the method or model used.
        labels (Optional[List[str]]): A list of class labels for the confusion matrix axes. If None, labels will not be
        set.

    Returns:
        None: The function saves the plot to the specified path and does not return any value.
    """

    fig, axis = plt.subplots(1, 3, figsize=(15, 5))

    for i, cm in enumerate(cm):
        ax = axis[i]
        sns.heatmap(cm, annot=True, fmt='.0f', xticklabels=labels, yticklabels=labels, ax=ax)
        ax.set_title('%s, %s, %s, %s' % (f"Phase {i + 1}", name_of_dataset, method, operation))
        ax.set_xlabel('Predicted')
        ax.set_ylabel('Actual')

    filename = os.path.join(path_to_plot, f"{name_of_dataset}_{method}_{operation}.png")
    plt.tight_layout()
    plt.savefig(filename, dpi=300)
    plt.close()
    gc.collect()


def reorder_metrics_lists(train_metrics, test_metrics, training_time_list: Optional = None) -> List:
    """
    Reorders and combines training and testing metrics into a single list of metrics.

    Args:
        train_metrics:
            A list of training metrics.
            Expected order: [train_acc, train_precision, train_recall, train_f1] or
            [train_acc, train_precision, train_recall, train_f1, training_time] if training_time_list is not provided.
        test_metrics:
            A list of testing metrics.
            Expected order: [test_acc, test_precision, test_recall, test_f1].
        training_time_list (Optional[List[Union[int, float]]]):
            An optional list of training times. If provided, its sum is used as the total training time.

    Returns:
        List: A list containing a single tuple with the reordered and combined metrics.

    Notes:
        - If `training_time_list` is provided, the function sums it and appends it to the combined metrics.
        - If `training_time_list` is not provided, the training time is taken from the `train_metrics` list.
    """

    if training_time_list is not None:
        training_time = sum(training_time_list)

        train_acc, train_precision, train_recall, train_f1, = (
            train_metrics[0], train_metrics[1], train_metrics[2], train_metrics[3]
        )
    else:
        train_acc, train_precision, train_recall, train_f1, training_time = (
            train_metrics[0], train_metrics[1], train_metrics[2], train_metrics[3], train_metrics[4]
        )

    test_acc, test_precision, test_recall, test_f1 = (
        test_metrics[0], test_metrics[1], test_metrics[2], test_metrics[3]
    )

    combined_metrics = [
        train_acc, test_acc,
        train_precision, test_precision,
        train_recall, test_recall,
        train_f1, test_f1,
        training_time
    ]

    return [tuple(combined_metrics)]


def save_log_to_txt(output_file: str, result: Any, operation: str) -> None:
    """
    Saves the best trial configuration and results to a text file based on the specified operation.

    Args:
        output_file (str): The path to the output text file where the log will be saved.
        result (Any): An object containing the results of the trials. This should have a `get_best_trial` method to
            retrieve the best trial based on the given criteria.
        operation (str): A string specifying the type of operation to log. Should be either "loss" or "accuracy".

    Returns:
        None: The function writes the log to the specified file and does not return any value.

    Raises:
        ValueError: If the provided operation is not "loss" or "accuracy".
    """

    original_stdout = sys.stdout

    with open(output_file, "w") as log_file:
        sys.stdout = log_file

        if operation == "loss":
            best_trial = result.get_best_trial("loss", "min", "last")
            print("Best trial config: {}".format(best_trial.config))
            print("Best trial final validation loss: {}".format(best_trial.last_result["loss"]))
            print("Best trial final validation accuracy: {}".format(best_trial.last_result["accuracy"]))
        elif operation == "accuracy":
            best_trial = result.get_best_trial("accuracy", "max", "last")
            print("Best trial config: {}".format(best_trial.config))
            print("Best trial final validation accuracy: {}".format(best_trial.last_result["accuracy"]))
        else:
            raise ValueError(f"Invalid operation: {operation}")

    sys.stdout = original_stdout

    logging.info(f"Saving log to {output_file}")


def setup_logger() -> logging.Logger:
    """
    Set up a colorized logger with the following log levels and colors:

    - DEBUG: Cyan
    - INFO: Green
    - WARNING: Yellow
    - ERROR: Red
    - CRITICAL: Red on a white background

    Returns:
        The configured logger instance.
    """

    logger = logging.getLogger()
    if logger.hasHandlers():
        return logger

    logger.setLevel(logging.INFO)

    formatter = colorlog.ColoredFormatter(
        "%(log_color)s%(levelname)-8s%(reset)s %(white)s%(message)s",
        log_colors={
            'DEBUG': 'cyan',
            'INFO': 'green',
            'WARNING': 'yellow',
            'ERROR': 'red',
            'CRITICAL': 'red,bg_white',
        })

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    return logger
